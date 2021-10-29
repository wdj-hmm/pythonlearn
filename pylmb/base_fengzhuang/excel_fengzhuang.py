#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl


class HandleExcel():
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        work = openpyxl.load_workbook(self.filename)
        shs = work[self.sheetname]
        res = list(shs.rows)
        title = [i.value for i in res[0]]
        cases = []
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_data(self, row, column, value):
        work = openpyxl.load_workbook(self.filename)
        shs = work[self.sheetname]
        # 写入数据到指定格子
        shs.cell(row=row, column=column, value=value)
        work.save(self.filename)


if __name__ == '__main__':
    c = HandleExcel(r'F:\pythonlearn\pylmb\cases.xlsx', 'login').read_data()
    print(c)
