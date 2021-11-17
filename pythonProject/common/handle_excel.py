#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl


class Handle_Excel():
    def __init__(self, filepath, sheetname):
        self.filename = filepath
        self.sheetname = sheetname

    def read_data(self):
        workplace = openpyxl.load_workbook(self.filename)
        shs = workplace[self.sheetname]
        res = list(shs.rows)
        title = [i.value for i in res[0]]
        cases = []
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_data(self, row, column, value):
        workplace = openpyxl.load_workbook(self.filename)
        shs = workplace[self.sheetname]
        shs.cell(row=row, column=column, value=value)
        workplace.save(self.filename)
